"""XLSX parser via openpyxl.

Emits one block per sheet-region. Very large sheets are chunked by row count
so no single block exceeds ~2K rows (which would otherwise blow up the
embedding window).
"""

from __future__ import annotations

import zipfile
from pathlib import Path

from eoditdeora.parsers.base import Block, ParsedDoc, ParseResult
from eoditdeora.parsers.registry import register
from eoditdeora.utils.paths_util import display_path

_ROW_CHUNK = 500


class XlsxParser:
    name = "xlsx_openpyxl"
    supported_extensions = ("xlsx", "xlsm")
    fidelity = 4

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower().lstrip(".") in self.supported_extensions

    def parse(self, path: Path, *, doc_id: str) -> ParseResult:
        if not path.exists():
            return ParseResult(
                doc=ParsedDoc(
                    doc_id=doc_id,
                    source_path=str(path),
                    source_path_display=display_path(path),
                    format="xlsx",
                    parser=self.name,
                    fidelity=self.fidelity,
                    parse_status="file_missing",
                    warnings=["file_missing"],
                )
            )
        if path.stat().st_size == 0:
            return ParseResult(
                doc=ParsedDoc(
                    doc_id=doc_id,
                    source_path=str(path),
                    source_path_display=display_path(path),
                    format="xlsx",
                    parser=self.name,
                    fidelity=self.fidelity,
                    parse_status="empty",
                    warnings=["empty_file"],
                )
            )

        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as e:
            return ParseResult(
                doc=ParsedDoc(
                    doc_id=doc_id,
                    source_path=str(path),
                    source_path_display=display_path(path),
                    format="xlsx",
                    parser=self.name,
                    fidelity=self.fidelity,
                    parse_status="parser_error",
                    warnings=[f"openpyxl_not_available: {e}"],
                )
            )

        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
        except zipfile.BadZipFile as e:
            return ParseResult(
                doc=ParsedDoc(
                    doc_id=doc_id,
                    source_path=str(path),
                    source_path_display=display_path(path),
                    format="xlsx",
                    parser=self.name,
                    fidelity=self.fidelity,
                    parse_status="invalid_format",
                    warnings=[f"xlsx_not_zip: {e}"],
                )
            )
        except Exception as e:  # noqa: BLE001
            return ParseResult(
                doc=ParsedDoc(
                    doc_id=doc_id,
                    source_path=str(path),
                    source_path_display=display_path(path),
                    format="xlsx",
                    parser=self.name,
                    fidelity=self.fidelity,
                    parse_status="invalid_format",
                    warnings=[f"xlsx_open_failed: {e}"],
                )
            )

        blocks: list[Block] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rendered = ["" if v is None else str(v) for v in row]
                if any(c.strip() for c in rendered):
                    rows.append(rendered)
                if len(rows) >= _ROW_CHUNK:
                    blocks.append(_rows_to_block(sheet_name, rows))
                    rows = []
            if rows:
                blocks.append(_rows_to_block(sheet_name, rows))

        try:
            props = wb.properties
            metadata = {
                "author": props.creator,
                "created_at": props.created.isoformat() if props.created else None,
                "modified_at": props.modified.isoformat() if props.modified else None,
                "application": "Microsoft Excel",
                "title": props.title,
            }
        except Exception:  # noqa: BLE001
            metadata = {}

        doc = ParsedDoc(
            doc_id=doc_id,
            source_path=str(path),
            source_path_display=display_path(path),
            format="xlsx",
            parser=self.name,
            fidelity=self.fidelity,
            blocks=blocks,
            metadata={k: v for k, v in metadata.items() if v is not None},
        )
        return ParseResult(doc=doc)


def _rows_to_block(sheet: str, rows: list[list[str]]) -> Block:
    text = "\n".join("\t".join(r) for r in rows)
    return Block(
        type="table",
        text=text,
        sheet=sheet,
        meta={"rows": len(rows), "cols": max((len(r) for r in rows), default=0)},
    )


register(XlsxParser())
